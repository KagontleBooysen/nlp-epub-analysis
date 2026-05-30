"""
build_rating_sheet.py
Creates three separate Excel rating sheets — one per rater.
The VADER score is hidden from raters (separate sheet).

Run AFTER extract_gold_standard.py:
    python build_rating_sheet.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              Protection)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# ── Load extracted sentences ───────────────────────────────────────────────────
if not os.path.exists('gold_standard_sentences.csv'):
    print("ERROR: gold_standard_sentences.csv not found.")
    print("Run extract_gold_standard.py first.")
    exit()

df = pd.read_csv('gold_standard_sentences.csv', encoding='utf-8')
print(f"Loaded {len(df)} sentences")

# ── Colour scheme ──────────────────────────────────────────────────────────────
GENRE_COLOURS = {
    "Gothic/Horror":        "FFE0E0",
    "Adventure":            "E0F0FF",
    "Romance":              "FFE0F5",
    "Science Fiction":      "E0FFE8",
    "Children's Fiction":   "FFFDE0",
    "Mystery/Crime":        "F0E0FF",
    "Philosophy/Essays":    "E8E8E8",
    "Historical/Political": "FFE8D0",
    "Poetry/Drama":         "D0F0FF",
    "Science/Non-fiction":  "E0FFD0",
}

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT     = Font(name="Calibri", size=10)
CENTRE        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin          = Side(style="thin", color="BBBBBB")
BORDER        = Border(top=thin, bottom=thin, left=thin, right=thin)

def make_rater_sheet(wb, rater_num, df):
    """Create one rating sheet for one rater."""
    ws = wb.create_sheet(title=f"Rater {rater_num}")

    # ── Instructions row ────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = (
        f"RATER {rater_num} — Gold Standard Annotation Sheet  |  "
        "Rate each sentence on a scale of −2 to +2  |  "
        "−2 = Very Negative   −1 = Negative   0 = Neutral   +1 = Positive   +2 = Very Positive  |  "
        "Enter your rating in the YELLOW column only. Do not discuss ratings with other raters."
    )
    ws["A1"].font      = Font(name="Calibri", bold=True, size=10, color="1F4E79")
    ws["A1"].alignment = LEFT_WRAP
    ws["A1"].fill      = PatternFill("solid", fgColor="E8F0FE")
    ws.row_dimensions[1].height = 40

    # ── Column headers ──────────────────────────────────────────────────────
    headers = ["ID", "Genre", "Title", "Sentence", "Word Count", f"Rater {rater_num} Score"]
    col_widths = [6, 18, 30, 80, 10, 14]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTRE
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 22

    # ── Data rows ────────────────────────────────────────────────────────────
    score_col = 6  # column F = rating column

    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        genre    = str(row['Genre'])
        sentence = str(row['Sentence'])
        word_ct  = len(sentence.split())
        genre_colour = GENRE_COLOURS.get(genre, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=genre_colour)
        rating_fill = PatternFill("solid", fgColor="FFFF99")  # yellow for rating col

        values = [
            int(row['Sentence_ID']),
            genre,
            str(row['Title'])[:50],
            sentence,
            word_ct,
            "",  # blank rating cell
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = BORDER
            cell.font      = BODY_FONT

            if col_idx == score_col:
                cell.fill      = rating_fill
                cell.alignment = CENTRE
                cell.protection = Protection(locked=False)
            elif col_idx == 4:  # sentence column
                cell.alignment = LEFT_WRAP
                cell.fill      = row_fill
            elif col_idx in [2, 3]:
                cell.fill      = row_fill
                cell.alignment = LEFT_WRAP
            else:
                cell.fill      = row_fill
                cell.alignment = CENTRE

        # Row height based on sentence length
        ws.row_dimensions[row_idx].height = max(30, min(80, word_ct * 1.4))

    # ── Data validation — only accept −2 to +2 ───────────────────────────────
    dv = DataValidation(
        type="whole",
        operator="between",
        formula1="-2",
        formula2="2",
        showErrorMessage=True,
        errorTitle="Invalid Score",
        error="Please enter a whole number between −2 and +2 only."
    )
    ws.add_data_validation(dv)
    last_row = 2 + len(df)
    dv.sqref = f"F3:F{last_row}"

    # ── Freeze panes ────────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Auto-filter ────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:F{last_row}"

    print(f"  ✓ Rater {rater_num} sheet created ({len(df)} rows)")
    return ws


def make_master_sheet(wb, df):
    """Create a master sheet with VADER scores — for researcher use only."""
    ws = wb.create_sheet(title="MASTER (Researcher Only)")

    ws.merge_cells("A1:H1")
    ws["A1"] = "MASTER SHEET — Contains VADER scores. Do NOT share this sheet with raters."
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ws["A1"].fill = PatternFill("solid", fgColor="C0392B")
    ws["A1"].alignment = CENTRE

    headers = ["ID", "Genre", "Title", "Sentence", "VADER Score",
               "Rater 1", "Rater 2", "Rater 3", "Mean Human", "Pearson r Note"]
    col_widths = [6, 18, 30, 80, 12, 12, 12, 12, 12, 20]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTRE
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        vals = [
            int(row['Sentence_ID']),
            str(row['Genre']),
            str(row['Title'])[:50],
            str(row['Sentence']),
            row.get('VADER_Compound', ''),
            '', '', '',  # rater columns to fill later
            f"=AVERAGE(F{row_idx}:H{row_idx})",
            "=CORREL($E$3:$E$1002,F$3:F$1002)"
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = BORDER
            cell.font      = BODY_FONT
            cell.alignment = CENTRE if col_idx != 4 else LEFT_WRAP

    ws.freeze_panes = "A3"
    print(f"  ✓ Master sheet created")


# ── Build workbooks ────────────────────────────────────────────────────────────

# One combined workbook with all three rater sheets + master
wb_combined = Workbook()
wb_combined.remove(wb_combined.active)  # remove default sheet

print("\nBuilding rating sheets...")
make_rater_sheet(wb_combined, 1, df)
make_rater_sheet(wb_combined, 2, df)
make_rater_sheet(wb_combined, 3, df)
make_master_sheet(wb_combined, df)

combined_path = 'gold_standard_MASTER.xlsx'
wb_combined.save(combined_path)
print(f"\n✓ Master workbook saved: {combined_path}")

# Three separate workbooks — one per rater (no VADER scores visible)
print("\nBuilding individual rater workbooks...")
for rater_num in [1, 2, 3]:
    wb = Workbook()
    wb.remove(wb.active)
    make_rater_sheet(wb, rater_num, df)
    path = f'gold_standard_Rater{rater_num}.xlsx'
    wb.save(path)
    print(f"  ✓ Saved: {path}")

print(f"""
{'='*55}
ALL FILES CREATED
{'='*55}
  gold_standard_sentences.csv    — raw extracted sentences
  gold_standard_MASTER.xlsx      — master sheet (researcher only)
  gold_standard_Rater1.xlsx      — send to Rater 1
  gold_standard_Rater2.xlsx      — send to Rater 2
  gold_standard_Rater3.xlsx      — keep for your own ratings

INSTRUCTIONS:
  1. You rate using gold_standard_Rater3.xlsx
  2. Send gold_standard_Rater1.xlsx to Rater 1
  3. Send gold_standard_Rater2.xlsx to Rater 2
  4. When all 3 are done, paste ratings into MASTER sheet
  5. Run compute_agreement.py to get Fleiss kappa and Pearson r
""")